# streamlit_app.py（ステップ1：入力・実行・結果表示）

import streamlit as st
import pandas as pd
from maps_scraper import get_google_maps_data
from openpyxl import load_workbook
import io

st.set_page_config(page_title="G-Maps企業情報抽出ツール", layout="wide")
st.title("📍 G-Maps企業情報抽出ツール")

# --- 入力欄 ---
region = st.text_input("🔍 検索する地域を入力してください（例：多治見市）")
industry = st.text_input("🏭 検索する業種を入力してください（例：製造工場）")

if region and industry:
    keyword = f"{region} {industry}"
    st.markdown(f"### 🔎 検索キーワード：**{keyword}**")
    
    if st.button("🔄 Googleマップから企業情報を取得"):
        with st.spinner("Googleマップを検索中..."):
            df = get_google_maps_data(keyword)

        st.success(f"✅ {len(df)} 件の企業情報を取得しました")
        st.dataframe(df)

        # Excelテンプレートへの出力
        try:
            template_path = "template.xlsx"
            wb = load_workbook(template_path)
            sheet = wb["入力マスター"]

            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                for cell in row[1:]:
                    cell.value = None

            for i, row in df.iterrows():
                sheet.cell(row=i+2, column=2, value=row["企業名"])
                sheet.cell(row=i+2, column=3, value=row["業種"])
                sheet.cell(row=i+2, column=4, value=row["住所"])
                sheet.cell(row=i+2, column=5, value=row["電話番号"])

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            st.download_button(
                label="📥 Excelでダウンロード",
                data=output,
                file_name=f"{region}_{industry}_企業リスト.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"❌ Excel出力中にエラーが発生しました：{e}")
else:
    st.info("👆 上記の地域・業種を入力してください")